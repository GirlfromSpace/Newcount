from PyQt5 import QtWidgets
import openpyxl


# загрузка таблицы из бд
def setup_table_data(self):
    path = "./database.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    self.ui_main.tableWidget.setRowCount(sheet.max_row)
    self.ui_main.tableWidget.setColumnCount(sheet.max_column)
    list_values = list(sheet.values)
    # загрузка данных в виджет
    self.ui_main.tableWidget.setHorizontalHeaderLabels(list_values[0])
    row_index = 0
    for i in list_values[1:]:
        col_index = 0
        for value in i:
            self.ui_main.tableWidget.setItem(row_index, col_index, QtWidgets.QTableWidgetItem(str(value)))
            col_index += 1
        row_index += 1